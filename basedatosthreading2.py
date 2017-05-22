from threading import Thread
import time
import warnings
from libs import communications, filterings, graphics, datalog, normalizado
import openpyxl


warnings.filterwarnings("ignore", category=DeprecationWarning)
communication = communications.CommunicationManager()
filtering = filterings.FilteringManager()
graphic = graphics.GraphicsManager()
report = datalog.DatalogManager()
filtros = normalizado.Normalizador()
communication.open_serial_port()
graphic.set_plot_parameters()


x = []
y = []
z = []
digitos_prediccion=[]

def adquieredatos():
    tiempo_inicial_adquisicion = time.time()
    while(len(digitos_prediccion)) <= 300:
        communication.send_data_request()
        time.sleep(0.25)
        if (len(digitos_prediccion)%30)==0 and len(digitos_prediccion)!=0:
            time.sleep(1)
            print('HAZ MOVIMIENTO')
            time.sleep(0.2)


def lee_y_escribe(index):
    contador=-1
    tiempo_inicial_lectura=time.time()
    n_iteraciones = 0
    fila=0
    etiqueta = 1
    while (len(digitos_prediccion)) <= 150:
        bytestoread, inbyte = communication.read_data()

        if (bytestoread == 7) or (bytestoread == 14):
            contador+=1
            print('Thread 2-inbyte: ' + str(inbyte) + ' length inbyte ' + str(len(inbyte)) +' bits to read '+str(bytestoread))
            x.append((inbyte[bytestoread - 3]))
            y.append((inbyte[bytestoread - 2]))
            z.append((inbyte[bytestoread - 1]))
            print('contador de thread 2 ' + str(contador))
            filtering.filter_acceleration(x, contador)
            filtering.filter_acceleration(y, contador)
            filtering.filter_acceleration(z, contador)
            digitos_prediccion.append(x[contador])
            digitos_prediccion.append(y[contador])
            digitos_prediccion.append(z[contador])
            print('Longitud de digitos prediccion: ' + str(len(digitos_prediccion)))
            if (len(digitos_prediccion)%30)==0 and len(digitos_prediccion)!=0:
                w = len(digitos_prediccion)/30
                print('W ='+str(w))
                index_inicial = int((w-1)*30)
                print(index_inicial)
                index_final = int((w*30)-1)
                print(index_final)
                digitos_muestra = digitos_prediccion[index_inicial:index_final]
                digitos_muestra.append(index)
                doc = openpyxl.load_workbook('datos_muestra.xlsx')
                hoja = doc.get_sheet_by_name('Hoja1')
                ws=doc.active
                ws.append(digitos_muestra)
                doc.save('datos_muestra.xlsx')

                time.sleep(0.7)
        time.sleep(0.25)
def lee_y_escribe2(index):
    contador=-1
    tiempo_inicial_lectura=time.time()
    n_iteraciones = 0
    fila=0
    etiqueta = 1
    while (len(digitos_prediccion)) <= 300:
        bytestoread, inbyte = communication.read_data()

        if (bytestoread == 7) or (bytestoread == 14):
            contador+=1
            print('Thread 2-inbyte: ' + str(inbyte) + ' length inbyte ' + str(len(inbyte)) +' bits to read '+str(bytestoread))
            x.append((inbyte[bytestoread - 3]))
            y.append((inbyte[bytestoread - 2]))
            z.append((inbyte[bytestoread - 1]))
            print('contador de thread 2 ' + str(contador))
            filtering.filter_acceleration(x, contador)
            filtering.filter_acceleration(y, contador)
            filtering.filter_acceleration(z, contador)
            digitos_prediccion.append(x[contador])
            digitos_prediccion.append(y[contador])
            digitos_prediccion.append(z[contador])
            print('Longitud de digitos prediccion: ' + str(len(digitos_prediccion)))
            if (len(digitos_prediccion)%30)==0 and len(digitos_prediccion)!=0:
                w = len(digitos_prediccion)/30
                print('W ='+str(w))
                index_inicial = int((w-1)*30)
                print(index_inicial)
                index_final = int((w*30)-1)
                print(index_final)
                digitos_muestra = digitos_prediccion[index_inicial:index_final]
                digitos_muestra.append(index)
                doc = openpyxl.load_workbook('datos_muestra.xlsx')
                hoja = doc.get_sheet_by_name('Hoja1')
                ws=doc.active
                ws.append(digitos_muestra)
                doc.save('datos_muestra.xlsx')

                time.sleep(0.7)
        time.sleep(0.25)
def lee_y_escribe3(index):
    contador=-1
    tiempo_inicial_lectura=time.time()
    n_iteraciones = 0
    fila=0
    etiqueta = 1
    while (len(digitos_prediccion)) <= 450:
        bytestoread, inbyte = communication.read_data()

        if (bytestoread == 7) or (bytestoread == 14):
            contador+=1
            print('Thread 2-inbyte: ' + str(inbyte) + ' length inbyte ' + str(len(inbyte)) +' bits to read '+str(bytestoread))
            x.append((inbyte[bytestoread - 3]))
            y.append((inbyte[bytestoread - 2]))
            z.append((inbyte[bytestoread - 1]))
            print('contador de thread 2 ' + str(contador))
            filtering.filter_acceleration(x, contador)
            filtering.filter_acceleration(y, contador)
            filtering.filter_acceleration(z, contador)
            digitos_prediccion.append(x[contador])
            digitos_prediccion.append(y[contador])
            digitos_prediccion.append(z[contador])
            print('Longitud de digitos prediccion: ' + str(len(digitos_prediccion)))
            if (len(digitos_prediccion)%30)==0 and len(digitos_prediccion)!=0:
                w = len(digitos_prediccion)/30
                print('W ='+str(w))
                index_inicial = int((w-1)*30)
                print(index_inicial)
                index_final = int((w*30)-1)
                print(index_final)
                digitos_muestra = digitos_prediccion[index_inicial:index_final]
                digitos_muestra.append(index)
                doc = openpyxl.load_workbook('datos_muestra.xlsx')
                hoja = doc.get_sheet_by_name('Hoja1')
                ws=doc.active
                ws.append(digitos_muestra)
                doc.save('datos_muestra.xlsx')

                time.sleep(0.7)
        time.sleep(0.25)
def lee_y_escribe4(index):
    contador=-1
    tiempo_inicial_lectura=time.time()
    n_iteraciones = 0
    fila=0
    etiqueta = 1
    while (len(digitos_prediccion)) <= 600:
        bytestoread, inbyte = communication.read_data()

        if (bytestoread == 7) or (bytestoread == 14):
            contador+=1
            print('Thread 2-inbyte: ' + str(inbyte) + ' length inbyte ' + str(len(inbyte)) +' bits to read '+str(bytestoread))
            x.append((inbyte[bytestoread - 3]))
            y.append((inbyte[bytestoread - 2]))
            z.append((inbyte[bytestoread - 1]))
            print('contador de thread 2 ' + str(contador))
            filtering.filter_acceleration(x, contador)
            filtering.filter_acceleration(y, contador)
            filtering.filter_acceleration(z, contador)
            digitos_prediccion.append(x[contador])
            digitos_prediccion.append(y[contador])
            digitos_prediccion.append(z[contador])
            print('Longitud de digitos prediccion: ' + str(len(digitos_prediccion)))
            if (len(digitos_prediccion)%30)==0 and len(digitos_prediccion)!=0:
                w = len(digitos_prediccion)/30
                print('W ='+str(w))
                index_inicial = int((w-1)*30)
                print(index_inicial)
                index_final = int((w*30)-1)
                print(index_final)
                digitos_muestra = digitos_prediccion[index_inicial:index_final]
                digitos_muestra.append(index)
                doc = openpyxl.load_workbook('datos_muestra.xlsx')
                hoja = doc.get_sheet_by_name('Hoja1')
                ws=doc.active
                ws.append(digitos_muestra)
                doc.save('datos_muestra.xlsx')

                time.sleep(0.7)
        time.sleep(0.25)


peticion_aceleracion = Thread(target=adquieredatos)
lectura_almacenado1 = Thread(target=lee_y_escribe,args=(0,))#
lectura_almacenado2= Thread(target=lee_y_escribe2,args=(1,))#
lectura_almacenado3 = Thread(target=lee_y_escribe3,args=(2,))
lectura_almacenado4 = Thread(target=lee_y_escribe4,args=(3,))
peticion_aceleracion.start()
print('Movimientos aleatorios')
lectura_almacenado1.start()
print(lectura_almacenado1.isAlive())
print(lectura_almacenado2.isAlive())
time.sleep(20)
print('Movimientos hacia arriba')
time.sleep(0.2)
lectura_almacenado2.start()
time.sleep(20)
print('Movimientos hacia abajo')
time.sleep(0.2)
lectura_almacenado3.start()
time.sleep(20)
print('Movimientos hacia derecha')
time.sleep(0.2)
lectura_almacenado4.start()
